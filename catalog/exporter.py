"""Выгрузка собранных товаров в Excel и JSON."""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="2F3640")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True)

COLUMNS = [
    ("name", "Название", 44),
    ("category", "Категория", 18),
    ("price", "Цена", 12),
    ("currency", "Валюта", 9),
    ("rating", "Рейтинг", 10),
    ("in_stock", "На складе", 11),
    ("reviews", "Отзывов", 10),
    ("upc", "Артикул", 20),
    ("availability", "Наличие", 24),
    ("description", "Описание", 60),
    ("url", "Ссылка", 46),
]


def style_header(worksheet, width_by_column: dict[int, int]) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 26
    for index, width in width_by_column.items():
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.freeze_panes = "A2"


def write_products(worksheet, products: list) -> None:
    worksheet.append([title for _, title, _ in COLUMNS])

    for product in products:
        data = product.as_dict()
        row = []
        for key, _, _ in COLUMNS:
            value = data.get(key)
            # Описание в ячейке Excel режем: длинные тексты ломают чтение таблицы.
            if key == "description" and isinstance(value, str) and len(value) > 300:
                value = value[:297] + "…"
            row.append(value)
        worksheet.append(row)

    style_header(worksheet, {i: width for i, (_, _, width) in enumerate(COLUMNS, start=1)})

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"

    worksheet.auto_filter.ref = worksheet.dimensions


def write_summary(worksheet, products: list) -> None:
    by_category: dict[str, list] = {}
    for product in products:
        by_category.setdefault(product.category or "без категории", []).append(product)

    worksheet.append(["Категория", "Товаров", "Средняя цена", "Мин", "Макс"])

    for name in sorted(by_category):
        items = by_category[name]
        prices = [p.price for p in items if isinstance(p.price, (int, float))]
        worksheet.append([
            name,
            len(items),
            round(sum(prices) / len(prices), 2) if prices else None,
            min(prices) if prices else None,
            max(prices) if prices else None,
        ])

    all_prices = [p.price for p in products if isinstance(p.price, (int, float))]
    ratings = [p.rating for p in products if isinstance(p.rating, int)]

    worksheet.append([])
    worksheet.append(["Всего товаров", len(products)])
    worksheet.cell(row=worksheet.max_row, column=1).font = BOLD
    if all_prices:
        worksheet.append(["Средняя цена", round(sum(all_prices) / len(all_prices), 2)])
        worksheet.append(["Самый дешёвый", min(all_prices)])
        worksheet.append(["Самый дорогой", max(all_prices)])
    if ratings:
        worksheet.append(["Средний рейтинг", round(sum(ratings) / len(ratings), 2)])

    style_header(worksheet, {1: 26, 2: 12, 3: 14, 4: 10, 5: 10})

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"


def to_excel(products: list, path: Path) -> None:
    workbook = Workbook()
    write_products(workbook.active, products)
    workbook.active.title = "Товары"
    write_summary(workbook.create_sheet("Сводка"), products)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def to_json(products: list, path: Path, max_description: int | None = None) -> None:
    """max_description подрезает описания — нужно, когда файл едет в браузер."""
    path.parent.mkdir(parents=True, exist_ok=True)

    payload = []
    for product in products:
        data = product.as_dict()
        if max_description and len(data.get("description") or "") > max_description:
            data["description"] = data["description"][: max_description - 1] + "…"
        payload.append(data)

    compact = max_description is not None
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=None if compact else 2, separators=(",", ":") if compact else None),
        encoding="utf-8",
    )
